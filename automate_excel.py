import os
import re
import datetime
import logging
import openpyxl
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
import pandas as pd

# ------------------------------------------------------------
# Logging Configuration
# ------------------------------------------------------------
logging.basicConfig(
    level=logging.DEBUG,  # Change to INFO or ERROR if you want less verbosity
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

# ------------------------------------------------------------
# Constants / Paths
# ------------------------------------------------------------
TEMPLATE_PATH = r"template_daily_recap.xlsx"  # Adjust if stored elsewhere
TEMPLATE_SHEET_NAME = "Template"
TOTAL_SHEET_NAME = "Total"

# ------------------------------------------------------------
# Safe Cell Writing (for merged cells)
# ------------------------------------------------------------
def safe_set_cell(ws, cell_ref, value):
    """
    Safely sets the value of a cell. If the cell is merged, update the top-left cell.
    """
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left.value = value
                return
    else:
        cell.value = value

# ------------------------------------------------------------
# Clear Data from Copied Worksheet
# ------------------------------------------------------------
def clear_sheet_data(sheet, start_row=7):
    """
    Clears cell values from start_row to the end of the sheet.
    This prevents copying any sample data from the TEMPLATE.
    """
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

# ------------------------------------------------------------
# Prompt for Date Range (mm-dd-yyyy)
# ------------------------------------------------------------
def prompt_date_range():
    """
    Prompts for a start and end date in mm-dd-yyyy format.
    Returns (start_date, end_date) as date objects.
    """
    print("Enter the date range for your daily sheets (mm-dd-yyyy).")
    start_date_str = input("  Start date (mm-dd-yyyy): ").strip()
    end_date_str   = input("  End date   (mm-dd-yyyy): ").strip()

    # Convert user input to datetime.date
    try:
        start_date = datetime.datetime.strptime(start_date_str, "%m-%d-%Y").date()
        end_date   = datetime.datetime.strptime(end_date_str, "%m-%d-%Y").date()
    except ValueError:
        logging.error("Invalid date format. Please use mm-dd-yyyy.")
        exit(1)

    # Basic checks (you can remove or adjust these if you allow future dates)
    today = datetime.date.today()
    if start_date > today or end_date > today:
        logging.error("Future dates are not allowed in this example.")
        exit(1)
    if start_date > end_date:
        logging.error("Start date must not be later than end date.")
        exit(1)

    logging.info(f"Date range: {start_date} to {end_date}")
    return start_date, end_date

# ------------------------------------------------------------
# Prompt for File Paths (CSV/TXT)
# ------------------------------------------------------------
def prompt_file_paths():
    """
    Prompts for one or more CSV/TXT file paths (comma-separated) and verifies each exists.
    Returns a list of file paths.
    """
    file_paths_str = input("Enter the path(s) to the data file(s) (CSV/TXT, comma-separated): ").strip()
    file_paths = [fp.strip().strip('"') for fp in file_paths_str.split(",")]
    for fp in file_paths:
        if not os.path.exists(fp):
            logging.error(f"File not found: {fp}")
            exit(1)
    return file_paths

# ------------------------------------------------------------
# Prompt for Hourly Rate
# ------------------------------------------------------------
def prompt_rate():
    """
    Prompts for an hourly rate and returns it as float.
    """
    rate_str = input("Enter the hourly rate: ").strip()
    try:
        rate = float(rate_str)
    except ValueError:
        logging.error("Invalid rate. Please enter a numeric value.")
        exit(1)
    logging.info(f"Hourly rate: {rate}")
    return rate

# ------------------------------------------------------------
# Read CSV/TXT into DataFrame
# ------------------------------------------------------------
def read_csv_data(data_file):
    """
    Reads a CSV or TXT file into a Pandas DataFrame.
    If file ends with '.txt', assumes tab-delimited. Otherwise, comma-delimited.
    """
    try:
        if data_file.lower().endswith(".txt"):
            df = pd.read_csv(data_file, sep='\t')
        else:
            df = pd.read_csv(data_file)
        logging.info(f"Data file '{data_file}' read with {len(df)} rows.")
    except Exception as e:
        logging.error(f"Error reading data file '{data_file}': {e}")
        exit(1)
    return df

# ------------------------------------------------------------
# Combine multiple CSV/TXT files
# ------------------------------------------------------------
def combine_csv_data(file_paths):
    """
    Reads and concatenates multiple CSV/TXT files into a single DataFrame.
    """
    df_list = []
    for fp in file_paths:
        df = read_csv_data(fp)
        df_list.append(df)
    if df_list:
        combined_df = pd.concat(df_list, ignore_index=True)
    else:
        combined_df = pd.DataFrame()
    return combined_df

# ------------------------------------------------------------
# Create a list of date objects from start_date to end_date
# ------------------------------------------------------------
def create_date_list(start_date, end_date):
    """
    Creates a list of date objects from start_date to end_date (inclusive).
    """
    date_list = []
    current = start_date
    while current <= end_date:
        date_list.append(current)
        current += datetime.timedelta(days=1)
    return date_list

# ------------------------------------------------------------
# Filter DataFrame by the given date range (if 'Date' column exists)
# ------------------------------------------------------------
def filter_df_by_date(df, start_date, end_date):
    """
    If a 'Date' column exists, parses it as mm-dd-yyyy or auto-detect,
    filters the DataFrame by the given date range.
    Returns (filtered_df, has_date_column).
    """
    if 'Date' in df.columns:
        # Attempt to parse the 'Date' column
        # If your data is strictly mm-dd-yyyy, you can specify the format
        # e.g. df['Date'] = pd.to_datetime(df['Date'], format='%m-%d-%Y', errors='coerce')
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date

        original_count = len(df)
        df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]
        logging.info(f"Filtered from {original_count} rows to {len(df)} rows by 'Date' column.")
        return df, True
    else:
        logging.info("No 'Date' column found; applying tasks to every date in the range.")
        return df, False

# ------------------------------------------------------------
# Fill a daily sheet with data
# ------------------------------------------------------------
def fill_daily_sheet(sheet, date_obj, data_rows, is_dataframe, start_row=7, fallback_date=None):
    """
    Populates a daily sheet:
      - Cell B1 = the date in mm-dd-yyyy if data exists, else today's date.
      - If no data for that date and single-day range, also write fallback_date to B3.
      - Writes your typical headers in row 6, then data from row 7 onward.
    Returns the last row used.
    """
    if is_dataframe:
        # If the DataFrame has a 'Date' column, filter to date_obj
        if 'Date' in data_rows.columns:
            day_df = data_rows[data_rows['Date'] == date_obj]
        else:
            day_df = data_rows
        records = day_df.to_dict(orient='records')
    else:
        # If not a DataFrame, assume it's already a list of dict
        records = data_rows

    # If no records, fill B1 with today's date (or fallback)
    if not records:
        today_str = datetime.date.today().strftime("%m-%d-%Y")
        sheet["B1"] = today_str
        if fallback_date is not None:
            sheet["B3"] = fallback_date.strftime("%m-%d-%Y")
    else:
        # If we have data, B1 is the date for that sheet
        sheet["B1"] = date_obj.strftime("%m-%d-%Y")

    # Freeze the top rows/columns
    sheet.freeze_panes = sheet["A7"]

    # Write column headers in row 6
    headers = ["Number", "Daily Work Description", "Hr", "Min", "Complete", "Follow up", "Supervisor Comments"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=6, column=col_idx).value = header

    # Write data rows
    current_row = start_row
    for row_dict in records:
        sheet.cell(row=current_row, column=1).value = row_dict.get("Number", "")
        sheet.cell(row=current_row, column=2).value = row_dict.get("Daily Work Description", "")
        sheet.cell(row=current_row, column=3).value = row_dict.get("Hr", "")
        sheet.cell(row=current_row, column=4).value = row_dict.get("Min", "")

        # "Complete" cell with color-coded font
        complete_val = row_dict.get("Complete", "")
        complete_cell = sheet.cell(row=current_row, column=5)
        complete_cell.value = complete_val
        if isinstance(complete_val, str):
            if complete_val.lower() == "yes":
                complete_cell.font = Font(color="008000")  # green
            elif complete_val.lower() == "no":
                complete_cell.font = Font(color="FF0000")  # red

        sheet.cell(row=current_row, column=6).value = row_dict.get("Follow up", "")
        sheet.cell(row=current_row, column=7).value = row_dict.get("Supervisor Comments", "")
        current_row += 1

    last_row = current_row - 1
    logging.info(f"{date_obj}: Populated rows {start_row} to {last_row}.")
    return last_row

# ------------------------------------------------------------
# Update the "Total" sheet with daily info
# ------------------------------------------------------------
def update_total_sheet(total_sheet, daily_info, rate):
    """
    Fills the 'Total' sheet with summary info:
      - B3:E3 = headers: Date, Hour, Rate, Total Cost
      - from row 4 downward, each date gets a row
      - uses formulas referencing each daily sheet to sum hours+minutes
    """
    # Clear old data from row 4 downward
    for row in total_sheet.iter_rows(min_row=4, max_row=total_sheet.max_row):
        for cell in row:
            cell.value = None

    row_idx = 4
    for sheet_name, (start_row, last_row) in sorted(daily_info.items()):
        if last_row < start_row:
            # Means no actual data
            continue

        # hour_formula sums hours (col C) plus minutes (col D) / 60
        hour_formula = f"=SUM('{sheet_name}'!C{start_row}:C{last_row}) + (SUM('{sheet_name}'!D{start_row}:D{last_row})/60)"
        safe_set_cell(total_sheet, f"B{row_idx}", sheet_name)  # The date (sheet name)
        safe_set_cell(total_sheet, f"C{row_idx}", hour_formula)
        safe_set_cell(total_sheet, f"D{row_idx}", rate)
        safe_set_cell(total_sheet, f"E{row_idx}", f"=C{row_idx}*D{row_idx}")
        row_idx += 1

# ------------------------------------------------------------
# Create or update the "Total" sheet
# ------------------------------------------------------------
def create_or_update_total_sheet(wb, daily_info, rate):
    """
    Ensures a 'Total' sheet exists, writes summary headers in B3:E3,
    populates summary rows, and freezes rows above row 4.
    """
    if TOTAL_SHEET_NAME in wb.sheetnames:
        total_sheet = wb[TOTAL_SHEET_NAME]
    else:
        total_sheet = wb.create_sheet(TOTAL_SHEET_NAME)

    safe_set_cell(total_sheet, "B3", "Date")
    safe_set_cell(total_sheet, "C3", "Hour")
    safe_set_cell(total_sheet, "D3", "Rate")
    safe_set_cell(total_sheet, "E3", "Total Cost")

    update_total_sheet(total_sheet, daily_info, rate)
    total_sheet.freeze_panes = total_sheet["A4"]
    return total_sheet

# ------------------------------------------------------------
# Main Workflow
# ------------------------------------------------------------
def main():
    """
    Main workflow:
      1) Prompt for date range in mm-dd-yyyy
      2) Prompt for CSV/TXT files
      3) Prompt for hourly rate
      4) Load template workbook
      5) Combine all file data -> DataFrame
         - if 'Date' column, parse & filter to range
         - else apply entire dataset to each day
      6) For each date in [start_date..end_date], copy "Template" -> fill data
      7) Create/Update "Total" sheet
      8) Hide "Template" sheet
      9) Save as either "mm-dd-yyyy.xlsx" or "mm-dd-yyyy_to_mm-dd-yyyy.xlsx"
    """
    # 1) Prompt date range
    start_date, end_date = prompt_date_range()

    # 2) Prompt file paths
    file_paths = prompt_file_paths()

    # 3) Prompt rate
    rate = prompt_rate()

    # 4) Load template workbook
    if not os.path.exists(TEMPLATE_PATH):
        logging.error(f"Template file not found: {TEMPLATE_PATH}")
        exit(1)
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        logging.error(f"No sheet named '{TEMPLATE_SHEET_NAME}' in {TEMPLATE_PATH}")
        exit(1)

    # 5) Combine data from CSV/TXT
    combined_df = combine_csv_data(file_paths)
    combined_df, has_date = filter_df_by_date(combined_df, start_date, end_date)

    # Create a list of dates from start_date..end_date
    date_list = []
    current = start_date
    while current <= end_date:
        date_list.append(current)
        current += datetime.timedelta(days=1)

    # We’ll track the (start_row, last_row) used in each daily sheet
    daily_info = {}

    # 6) For each date, copy the template, fill data
    for day in date_list:
        sheet_name = day.strftime("%m-%d-%Y")  # name the sheet as mm-dd-yyyy
        new_sheet = wb.copy_worksheet(wb[TEMPLATE_SHEET_NAME])
        new_sheet.title = sheet_name

        # Clear any data from the template
        clear_sheet_data(new_sheet, start_row=7)

        # If we have a 'Date' column, filter to that day; else use entire combined_df
        # fill_daily_sheet does the final step
        fallback = day if (start_date == end_date) else None
        last_row = fill_daily_sheet(
            new_sheet,
            date_obj=day,
            data_rows=combined_df,  # pass the entire filtered DF
            is_dataframe=True,
            start_row=7,
            fallback_date=fallback
        )
        daily_info[sheet_name] = (7, last_row)

    # 7) Create/Update "Total" sheet
    create_or_update_total_sheet(wb, daily_info, rate)

    # 8) Hide the template sheet
    wb[TEMPLATE_SHEET_NAME].sheet_state = "hidden"

    # 9) Save the workbook
    if start_date == end_date:
        output_filename = f"{start_date.strftime('%m-%d-%Y')}.xlsx"
    else:
        output_filename = f"{start_date.strftime('%m-%d-%Y')}_to_{end_date.strftime('%m-%d-%Y')}.xlsx"

    try:
        wb.save(output_filename)
    except PermissionError as e:
        logging.error(f"Permission error saving '{output_filename}': {e}")
        exit(1)

    logging.info(f"Workbook '{output_filename}' created successfully.")

# ------------------------------------------------------------
# Entry Point
# ------------------------------------------------------------
if __name__ == "__main__":
    main()
